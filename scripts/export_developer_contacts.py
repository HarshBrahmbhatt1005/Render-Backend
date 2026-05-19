#!/usr/bin/env python3
"""
Export unique developer contacts from Builder Visit data to Excel.

Default output:
- Backend folder: Render-Backend-main/Render-Backend-main/developer_contacts.xlsx

Optional output target:
- Frontend folder (MIS-Intigration2-main) via --output-target frontend

Usage examples:
  python scripts/export_developer_contacts.py
  python scripts/export_developer_contacts.py --output-target frontend
  python scripts/export_developer_contacts.py --output-file developer_contacts_2026.xlsx
  python scripts/export_developer_contacts.py --mongo-uri "mongodb://localhost:27017/mydb"

Required environment (if --mongo-uri not provided):
- MONGO_URI in backend .env file
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Set, Tuple

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None

try:
    from pymongo import MongoClient
except ImportError:
    print("Missing dependency: pymongo")
    print("Install with: pip install pymongo")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("Missing dependency: openpyxl")
    print("Install with: pip install openpyxl")
    sys.exit(1)


def normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip())


def normalize_number(number: str) -> str:
    digits = re.sub(r"\D", "", number or "")
    if not digits:
        return ""
    if len(digits) > 10:
        digits = digits[-10:]
    return digits


def get_backend_root(script_path: Path) -> Path:
    return script_path.resolve().parents[1]


def get_frontend_root(backend_root: Path) -> Path:
    workspace_root = backend_root.parent.parent
    return workspace_root / "MIS-Intigration2-main"


def resolve_output_dir(backend_root: Path, output_target: str, custom_output_dir: Optional[str]) -> Path:
    if custom_output_dir:
        return Path(custom_output_dir).resolve()

    if output_target == "backend":
        return backend_root

    if output_target == "frontend":
        frontend_root = get_frontend_root(backend_root)
        return frontend_root

    return backend_root


def extract_contact_pairs(doc: Dict) -> Iterable[Tuple[str, str]]:
    # Most likely developer office person fields
    office_name = (doc.get("officePersonDetails") or "").strip()
    office_number = (doc.get("officePersonNumber") or "").strip()

    # Secondary fallback fields found in same cards/schema
    builder_name = (doc.get("builderName") or "").strip()
    builder_number = (doc.get("builderNumber") or "").strip()

    candidates = [
        (office_name, office_number),
        (builder_name, builder_number),
    ]

    for name, number in candidates:
        if not name or not number:
            continue
        yield name, number


def fetch_unique_contacts(mongo_uri: str, collection_name: str) -> List[Tuple[str, str]]:
    client = MongoClient(mongo_uri)

    try:
        db = client.get_default_database()
        if db is None:
            raise ValueError(
                "Could not infer database name from MONGO_URI. Use URI with DB name, e.g. mongodb://host:27017/DB_NAME"
            )

        collection = db[collection_name]
        projection = {
            "officePersonDetails": 1,
            "officePersonNumber": 1,
            "builderName": 1,
            "builderNumber": 1,
        }

        seen_keys: Set[Tuple[str, str]] = set()
        unique_rows: List[Tuple[str, str]] = []

        for doc in collection.find({}, projection):
            for raw_name, raw_number in extract_contact_pairs(doc):
                name = normalize_name(raw_name)
                number = normalize_number(raw_number)

                if not name or not number:
                    continue

                dedupe_key = (name.lower(), number)
                if dedupe_key in seen_keys:
                    continue

                seen_keys.add(dedupe_key)
                unique_rows.append((name, number))

        unique_rows.sort(key=lambda item: (item[0].lower(), item[1]))
        return unique_rows
    finally:
        client.close()


def write_excel(rows: List[Tuple[str, str]], output_file: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Developer Contacts"

    ws.append(["Developer Name", "Developer Number"])

    header_fill = PatternFill("solid", fgColor="FFF2CC")
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for name, number in rows:
        ws.append([name, number])

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22

    for row in ws.iter_rows(min_row=2, max_col=2):
        row[0].alignment = Alignment(vertical="center")
        row[1].alignment = Alignment(horizontal="center", vertical="center")

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export unique developer name/number pairs from Builder Visit data to Excel."
    )
    parser.add_argument("--mongo-uri", default=None, help="Mongo URI. If omitted, uses MONGO_URI from backend .env")
    parser.add_argument(
        "--collection",
        default="buildervisits",
        help="Mongo collection name (default: buildervisits)",
    )
    parser.add_argument(
        "--output-target",
        choices=["backend", "frontend"],
        default="backend",
        help="Where to place output file if --output-dir is not set (default: backend)",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Custom output directory path (overrides --output-target)",
    )
    parser.add_argument(
        "--output-file",
        default=None,
        help="Output file name (default: developer_contacts_<timestamp>.xlsx)",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    script_path = Path(__file__)
    backend_root = get_backend_root(script_path)

    env_file = backend_root / ".env"
    if load_dotenv is not None and env_file.exists():
        load_dotenv(env_file)

    mongo_uri = args.mongo_uri or os.getenv("MONGO_URI")
    if not mongo_uri:
        print("MONGO_URI not found. Set it in backend .env or pass --mongo-uri.")
        return 1

    output_dir = resolve_output_dir(backend_root, args.output_target, args.output_dir)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_name = args.output_file or f"developer_contacts_{timestamp}.xlsx"
    output_file = output_dir / output_name

    try:
        rows = fetch_unique_contacts(mongo_uri, args.collection)
    except Exception as exc:
        print(f"Failed to fetch data from MongoDB: {exc}")
        return 1

    if not rows:
        print("No developer name-number pairs found.")
        return 0

    try:
        write_excel(rows, output_file)
    except Exception as exc:
        print(f"Failed to write Excel file: {exc}")
        return 1

    print(f"Export complete. Rows written: {len(rows)}")
    print(f"File created: {output_file}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
